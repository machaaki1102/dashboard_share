import openpyxl
import streamlit
import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
#from openpyxl.drawing.image import Image
from copy import copy
import os
import tempfile
from PIL import Image as PILImage  # PillowのImageクラスをインポート
from openpyxl.drawing.image import Image as OpenpyxlImage

# タイトルを追加
st.title('パンフレット作成')

# カテゴリごとのリストを作成
fruits = ['いちご', 'ぶどう', 'りんご', 'バナナ', 'オレンジ', 'パイナップル', 'キウイ', 'マンゴー', 'もも', 'さくらんぼ']
sports = ['サッカー', 'バスケットボール', 'テニス', '野球', 'バレーボール', 'ラグビー', '卓球', 'バドミントン', 'ゴルフ', 'ホッケー']
fish = ['サケ', 'マグロ', 'カツオ', 'イワシ', 'サバ', 'アジ', 'ヒラメ', 'カレイ', 'タラ', 'イカ']

# ファイルパスを指定してExcelファイルを読み込む
@st.cache_data
def load_data(file_path):
    #df = pd.read_csv(file_path)  # 例: CSVファイルの読み込み
    df = pd.read_excel('銘柄データ_BB.xlsx')
    # '肥料名称' カラムから NaN を取り除く
    df = df.dropna(subset=['肥料名称'])
    return df

if st.button('Clear Cache'):
    st.cache_data.clear()

df = load_data('銘柄データ_BB.xlsx')


# '肥料名称' カラムから NaN を取り除く
#df = df.dropna(subset=['肥料名称'])

#肥料名称のリストをつくる。
fertilizer_names = df['肥料名称'].tolist()

# 選択されたアイテムのリストを作成
selected_fruits = []
selected_sports = []
selected_fish = []
selected_fertilizer = []

# 3つのカラムを作成
col1, col2, col3 = st.columns(3)

# 1列目にフルーツのチェックボックスを作成
#with col1:
#    st.header("BB")
#    for fruit in fruits:
#        if st.checkbox(fruit, key=fruit):
#            selected_fruits.append(fruit)

# 1列目にフルーツのチェックボックスを作成
with col1:
    st.header("BB")
    for fertilizer_name in fertilizer_names:
        if st.checkbox(fertilizer_name, key=fertilizer_name):
            selected_fertilizer.append(fertilizer_name)


# 2列目に球技のチェックボックスを作成
with col2:
    st.header("化成")
    for sport in sports:
        if st.checkbox(sport, key=sport):
            selected_sports.append(sport)

# 3列目に魚のチェックボックスを作成
with col3:
    st.header("液肥")
    for fish_item in fish:
        if st.checkbox(fish_item, key=fish_item):
            selected_fish.append(fish_item)

# 選択されたアイテムのリストを表示
#st.write('選択された肥料:', selected_fertilizer)
#st.write('選択された球技:', selected_sports)
#st.write('選択された魚:', selected_fish)

# 選択されたアイテムの数を表示
selected_fertilizer_count = len(selected_fertilizer)
#selected_sports_count = len(selected_sports)
#selected_fish_count = len(selected_fish)

#st.write(selected_fruits_count)
#st.write(selected_sports_count)
#st.write(selected_fish_count)

if st.button('開始する'):
    # ワークブックをロードする
    wb = openpyxl.load_workbook('bb_tem.xlsx')
    # ワークシートを選択する（シート名を指定する）
    ws = wb['BB_テンプレ']

    # 必要数
    count_number = selected_fertilizer_count  ###ここがチェックされた数字となる。
    #テンプレートを作るところ。
    m = count_number - 1  # ここでチェックをつけられた分だけコピーすることになる。0からカウント
    count = (m // 2)

    for i in range(0, count):
        row_count = 1
        col_count = 14
        col_offset = i * 13

        # コピー元の範囲（例: A1からM44）
        source_range = [[ws.cell(row=r, column=c) for c in range(1, 14)] for r in range(1, 45)]

        # コピー先の左上セル（例: N1）
        dest_start_cell = ws.cell(row=row_count, column=col_count + col_offset)

        def copy_cell(src_cell, dest_cell):
            dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = copy(src_cell.number_format)
                dest_cell.protection = copy(src_cell.protection)
                dest_cell.alignment = copy(src_cell.alignment)

        # コピー元範囲の行数と列数を取得する
        row_count = len(source_range)
        col_count = len(source_range[0])

        # コピー元範囲をループしてコピー先にペーストする
        for i in range(row_count):
            for j in range(col_count):
                src_cell = source_range[i][j]
                dest_cell = ws.cell(row=dest_start_cell.row + i, column=dest_start_cell.column + j)
                copy_cell(src_cell, dest_cell)

        # 指定された列幅にコピー元とコピー先の列幅を設定する
        specified_widths = [1, 5.67, 8.42, 5.67, 4, 0.84, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08, 8.08]

        # コピー元の列幅を設定する
        for idx, width in enumerate(specified_widths, start=source_range[0][0].column):
            col_letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

        # コピー先の列幅を設定する
        for idx, width in enumerate(specified_widths, start=dest_start_cell.column):
            col_letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

    # いらないところを消す
    number = m + 1  # mが0からカウントとなるため、+1とする
    kesu_offset = (number // 2) * 13

    # 奇数の時のみ実行する
    if number % 2 != 0:
        # A1:M44 の範囲のセルをループする
        for row in ws.iter_rows(min_row=24, max_row=42, min_col=1 + kesu_offset, max_col=13 + kesu_offset):
            for cell in row:
                # セルの文字を消す
                cell.value = None

                # セルの罫線を消す
                cell.border = Border()

                # セルの背景色を消す (デフォルトは白)
                cell.fill = PatternFill(fill_type=None)

                # セルのフォントスタイルをデフォルトにリセット
                cell.font = Font()

    #ここからデータを入れるところ
    # 選択されたデータ数分入力する。リスト分0スタートなので、+1とする。
    mm = count_number

    i = 0
    # 各肥料名についてループ
    for fertilizer in selected_fertilizer:
        selected_row = df[df['肥料名称'] == fertilizer]
        
    #   for i in range(0, mm):
        row_offset = (i % 2) * 20
        col_offset = (i // 2) * 13
        
        # cはセル番地でH5を取得、NPK、速攻性、被覆尿素まで入れる
        n_base_row = 5
        n_base_column = 8
        
        name = ws.cell(row=n_base_row + row_offset - 1, column=n_base_column + col_offset - 7)
        name.value = selected_row['肥料名称'].values[0]
        
        # N入力
        n = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset)
        n.value = selected_row['N'].values[0]
        # P入力
        p = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset + 1)
        p.value = selected_row['P'].values[0]
        # K入力
        k = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset + 2)
        k.value = selected_row['K'].values[0]
        # 速攻性N
        s_n = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset + 3)
        s_n.value = selected_row['速効性'].values[0]
        # 被覆尿素
        h_n = ws.cell(row=n_base_row + row_offset, column=n_base_column + col_offset + 4)
        h_n.value = selected_row['被覆尿素'].values[0]
        # その他
        ano = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset)
        ano.value = selected_row['容量②'].values[0]
        # 栽培適正
        tekisei = ws.cell(row=n_base_row + row_offset + 12, column=n_base_column + col_offset - 6)
        tekisei.value = selected_row['栽培適正'].values[0]
        # 品種
        hinshu = ws.cell(row=n_base_row + row_offset + 13, column=n_base_column + col_offset - 6)
        hinshu.value = selected_row['品種'].values[0]
        # 特徴①
        tokuchou_1 = ws.cell(row=n_base_row + row_offset + 4, column=n_base_column + col_offset - 1)
        tokuchou_1.value = selected_row['特徴①'].values[0]
        # 特徴②
        tokuchou_2 = ws.cell(row=n_base_row + row_offset + 5, column=n_base_column + col_offset - 1)
        tokuchou_2.value = selected_row['特徴②'].values[0]
        # 特徴③
        tokuchou_3 = ws.cell(row=n_base_row + row_offset + 6, column=n_base_column + col_offset - 1)
        tokuchou_3.value = selected_row['特徴③'].values[0]
        # 特徴④
        tokuchou_4 = ws.cell(row=n_base_row + row_offset + 7, column=n_base_column + col_offset - 1)
        tokuchou_4.value = selected_row['特徴④'].values[0]
        # 特徴⑤
        tokuchou_5 = ws.cell(row=n_base_row + row_offset + 8, column=n_base_column + col_offset - 1)
        tokuchou_5.value = selected_row['特徴⑤'].values[0]
        
        
        # 容器,肥効曲線の画像を貼り付ける
        # スクリプトのディレクトリを取得
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # 絶対パスを生成
        img_path = os.path.join(script_dir, '容器', f'{fertilizer}.jpg')
        img_path2 = os.path.join(script_dir, '肥効曲線', f'{fertilizer}.jpg')

        # ファイルの存在を確認
        if os.path.exists(img_path):
            
            # Pillowで画像を開く
            original_img = PILImage.open(img_path)

            # 画像のリサイズ
            new_size = (190, 257)  # 新しいサイズを指定
            resized_img = original_img.resize(new_size)
          
            # 一時的なファイルを作成
            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_file:
                temp_path = tmp_file.name
                resized_img.save(temp_path)
#                st.write(f"Image temporarily saved at {temp_path}")

            # openpyxlのImageクラスでリサイズされた画像を読み込む
            img = OpenpyxlImage(temp_path)
            # Excelのセルに画像を貼り付ける位置を指定
            cell_address = ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset - 6).coordinate
            img.anchor = cell_address

            # 画像をシートに追加
            ws.add_image(img)
            # 一時ファイルを削除
            #os.remove(temp_path) 
        
        else:
            # ファイルが存在しない場合は何もしない
            pass

#====================================
        if os.path.exists(img_path2):
            
            # Pillowで画像を開く
            original_img2 = PILImage.open(img_path2)

            # 画像のリサイズ
            new_size2 = (440, 170)  # 新しいサイズを指定
            resized_img2 = original_img2.resize(new_size2)
          
            # 一時的なファイルを作成
            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_file2:
                temp_path2 = tmp_file2.name
                resized_img2.save(temp_path2)
#                st.write(f"Image temporarily saved at {temp_path}")

            # openpyxlのImageクラスでリサイズされた画像を読み込む
            img2 = OpenpyxlImage(temp_path2)
            # Excelのセルに画像を貼り付ける位置を指定
            cell_address2 = ws.cell(row=n_base_row + row_offset + 10, column=n_base_column + col_offset - 1).coordinate
 
#            ws.cell(row=n_base_row + row_offset + 1, column=n_base_column + col_offset - 6).coordinate
            img2.anchor = cell_address2

            # 画像をシートに追加
            ws.add_image(img2)
            # 一時ファイルを削除
            #os.remove(temp_path) 
        
        else:
            # ファイルが存在しない場合は何もしない
            pass
#=======
        i = i + 1
    # 変更を保存する
    wb.save('bb_tem_finish.xlsx')

# ファイルをバイナリモードで開く
with open('bb_tem_finish.xlsx', 'rb') as file:
    excel_data = file.read()

# ダウンロードボタンの作成
st.download_button(
    label="Download Excel File",  # ボタンのラベル
    data=excel_data,  # ダウンロードするデータ
    file_name='bb_tem_finish.xlsx',  # ダウンロード時のファイル名
    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIMEタイプを指定
)